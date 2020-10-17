import tkinter as tk
from tkinter import filedialog
import pandas as pd
import os.path


root = tk.Tk()

canvas1 = tk.Canvas(root, width=300, height=300, bg='lightsteelblue')
canvas1.pack()


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to
                       `DataFrame.to_excel()` [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        # print(startrow)
        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    # print(df)
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def getExcel():
    global df
    import_file_path = filedialog.askopenfilename()
    df = pd.read_excel(import_file_path)
    # print (df)
    it_df = pd.read_excel(import_file_path, sheet_name='IT data')
    network_df = pd.read_excel(import_file_path,
                               sheet_name='Network data')
    # sort by Priority
    itdata_cat12 = it_df[it_df.Priority.isin(["Category1", "Category2"])]
    itdata_cat34 = it_df[it_df.Priority.isin(["Category3", "Category4"])]
    networkdata_cat12 = network_df[network_df.Priority.isin(
        ["Category1", "Category2"])]
    networkdata_cat34 = network_df[network_df.Priority.isin(
        ["P3:Medium", "P4:Minor"])]

    # print (itdata_cat12)
    # path to files to be written
    write_file_path = "/Users/"
    itdata_cat12_file = os.path.join( write_file_path, "/Users/ITCat12.xlsx"")
    print(itdata_cat12_file)
    itdata_cat34_file = os.path.join( write_file_path,  "/Users/ITCat34.xlsx")
    print(itdata_cat34_file)
    networkdata_cat12_file = os.path.join( write_file_path,  "/Users/NetworkCat12.xlsx")
    print(nnetworkdata_cat12_file)
    networkdata_cat34_file = os.path.join( write_file_path,  "/Users/NetworkCat34.xlsx")
    print(networkdata_cat34_file)
    # IT data
    try:

        append_df_to_excel(itdata_cat12_file, itdata_cat12, sheet_name='Sheet1',
                           truncate_sheet=False, header=None, index=False)
        append_df_to_excel(itdata_cat34_file, itdata_cat34, sheet_name='Sheet1',
                           truncate_sheet=False, header=None, index=False)
        append_df_to_excel(networkdata_cat12_file, networkdata_cat12, sheet_name='Sheet1',
                           truncate_sheet=False, header=None, index=False)
        append_df_to_excel(networkdata_cat34_file, networkdata_cat34, sheet_name='Sheet1',
                           truncate_sheet=False, header=None, index=False)
                       #itdata_cat12.to_excel(writer, sheet_name='Sheet1')
        #writer.save()
        print("done")
    except:
        print("error")

browseButton_Excel = tk.Button(text='Import Excel File', command=getExcel,
                               bg='green', fg='white', font=('helvetica', 12, 'bold'))
canvas1.create_window(150, 150, window=browseButton_Excel)


root.mainloop()
